from flask import Flask, jsonify, logging, render_template, request
from flask_socketio import SocketIO, emit
from flask_cors import CORS
import paho.mqtt.client as mqtt
import sqlite3
import datetime
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from werkzeug.security import generate_password_hash, check_password_hash
from flask import session, redirect, url_for
from flask import Flask, render_template, request, send_file
from flask_socketio import SocketIO
from flask_cors import CORS
import paho.mqtt.client as mqtt
import sqlite3
import datetime
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from werkzeug.security import generate_password_hash, check_password_hash
from flask import session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from collections import defaultdict
import logging


# ---------------- FLASK ----------------
app = Flask(__name__)
app.config['SECRET_KEY'] = 'secret!'
socketio = SocketIO(app, cors_allowed_origins="*", manage_session=False)
CORS(app)

# ---------------- MQTT ----------------
broker = "evoluzn.org"
port = 18889
USERNAME = "evzin_led"
PASSWORD = "63I9YhMaXpa49Eb"

mqttc = mqtt.Client()

DB_FILE = "device_data110.db"

# ---------------- DATABASE SETUP ----------------
conn = sqlite3.connect(DB_FILE)
cursor = conn.cursor()

# meter data table
cursor.execute("""
CREATE TABLE IF NOT EXISTS meter_data (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    deviceID TEXT,

    vol1 REAL,
    cur1 REAL,
    pow1 REAL,
    energy1 REAL,
    freq1 REAL,
    pf1 REAL,

    vol2 REAL,
    cur2 REAL,
    pow2 REAL,
    energy2 REAL,
    freq2 REAL,
    pf2 REAL,

    vol3 REAL,
    cur3 REAL,
    pow3 REAL,
    energy3 REAL,
    freq3 REAL,
    pf3 REAL,

    load_status INTEGER DEFAULT 0,

    inserttimestamp TEXT
)
""")


# Threshold 
cursor.execute("""
CREATE TABLE IF NOT EXISTS device_thresholds (
    device_id TEXT PRIMARY KEY,
    vol_min   REAL,
    vol_max   REAL,
    cur_min   REAL,
    cur_max   REAL,
    updated_at TEXT
)
""")

# alerts
cursor.execute("""
CREATE TABLE IF NOT EXISTS alerts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    deviceID TEXT,
    alert_type TEXT,
    phase INTEGER,
    message TEXT,
    timestamp TEXT
)
""")

# device status
cursor.execute("""
CREATE TABLE IF NOT EXISTS device_status (
    device_id TEXT PRIMARY KEY,
    status TEXT,
    last_seen TEXT
)
""")

# ---- USERS TABLE ----
cursor.execute("""
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    email TEXT UNIQUE,
    password TEXT
)
""")

# ── Table: device time settings ──────────────────────────────────────────────
cursor.execute("""
    CREATE TABLE IF NOT EXISTS device_time_settings (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        deviceID    TEXT    NOT NULL,
        hour        INTEGER NOT NULL,
        minute      INTEGER NOT NULL,
        day         INTEGER NOT NULL,
        month       INTEGER NOT NULL,
        year        INTEGER NOT NULL,
        set_at      TEXT    DEFAULT (datetime('now','localtime'))
    )
""")

# ── Table: device scheduling ──────────────────────────────────────────────────
cursor.execute("""
    CREATE TABLE IF NOT EXISTS device_scheduling (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        deviceID        TEXT    NOT NULL,
        is_enabled      INTEGER NOT NULL DEFAULT 0,   -- 1=ON, 0=OFF
        on_hour         INTEGER,
        on_minute       INTEGER,
        off_hour        INTEGER,
        off_minute      INTEGER,
        set_at          TEXT    DEFAULT (datetime('now','localtime'))
    )
""")

conn.commit()
conn.close()

# ---------------- GLOBALS ----------------
TARGET_DEVICE = "3Phase862733079332931"

# ---------------- MQTT CALLBACKS ----------------
def on_connect(client, userdata, flags, rc):
    if rc == 0:
        print("✅ MQTT CONNECTED")
        client.subscribe("+/status", qos=1)
    else:
        print("❌ MQTT FAILED")

def on_disconnect(client, userdata, rc):
    print("⚠ MQTT DISCONNECTED")


# ---------------- MQTT MESSAGE HANDLER ----------------
def on_message(client, userdata, message):
    try:
        payload = message.payload.decode().strip()
        topic = message.topic

         # 🚫 Ignore control messages (FIX)
        if topic.endswith("/control"):
            print(f"🚫 Ignored control message: {topic}")
            return

        print(f"📩 MQTT Received | Topic: {topic} | Payload: {payload}")
         
         # Only process status topics
        if not topic.endswith("/status"):
            print("⚠ Skipping non-status topic")
            return
        
        # Remove {}
        payload = payload.strip("{}")

        data = payload.split(":")

        if len(data) < 20:
            print("⚠ Invalid meter payload length")
            return

        deviceID = data[1]

        if deviceID != TARGET_DEVICE:
            return

        try:
            vol1 = float(data[2])
            cur1 = float(data[3])
            pow1 = float(data[4])
            energy1 = float(data[5])
            freq1 = float(data[6])
            pf1 = float(data[7])

            vol2 = float(data[8])
            cur2 = float(data[9])
            pow2 = float(data[10])
            energy2 = float(data[11])
            freq2 = float(data[12])
            pf2 = float(data[13])

            vol3 = float(data[14])
            cur3 = float(data[15])
            pow3 = float(data[16])
            energy3 = float(data[17])
            freq3 = float(data[18])
            pf3 = float(data[19])

        except ValueError:
            print("⚠ Invalid numeric data")
            return

        now = datetime.datetime.now()
        ts = now.strftime('%Y-%m-%d %H:%M:%S')
        time_label = now.strftime("%H:%M")

        with sqlite3.connect(DB_FILE) as conn:
            cur = conn.cursor()

            cur.execute("""
            INSERT OR REPLACE INTO device_status
            (device_id, status, last_seen)
            VALUES (?, ?, ?)
            """, (deviceID, "online", ts))

            cur.execute("""
            INSERT INTO meter_data (
                deviceID,
                vol1, cur1, pow1, energy1, freq1, pf1,
                vol2, cur2, pow2, energy2, freq2, pf2,
                vol3, cur3, pow3, energy3, freq3, pf3,
                inserttimestamp
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                deviceID,
                vol1, cur1, pow1, energy1, freq1, pf1,
                vol2, cur2, pow2, energy2, freq2, pf2,
                vol3, cur3, pow3, energy3, freq3, pf3,
                ts
            ))

            conn.commit()
                # ── Threshold check ─────────────────────────────────────────────────────────
        with sqlite3.connect(DB_FILE) as tconn:
            tconn.row_factory = sqlite3.Row
            thresh = tconn.execute(
                "SELECT * FROM device_thresholds WHERE device_id=?", (deviceID,)
            ).fetchone()

            if thresh:
                phases = [
                    ("R", 1, vol1, cur1),
                    ("Y", 2, vol2, cur2),
                    ("B", 3, vol3, cur3),
                ]
                alerts_to_insert = []

                for phase_name, phase_num, vol, cur_val in phases:

                    if thresh["vol_min"] is not None:
                        if vol < thresh["vol_min"] or vol > thresh["vol_max"]:
                            direction = "below min" if vol < thresh["vol_min"] else "above max"
                            alerts_to_insert.append((
                                deviceID, "VOLTAGE", phase_num,
                                f"Phase {phase_name}: Voltage {vol}V is {direction} "
                                f"(Threshold: {thresh['vol_min']}–{thresh['vol_max']}V)",
                                ts
                            ))

                    if thresh["cur_min"] is not None:
                        if cur_val < thresh["cur_min"] or cur_val > thresh["cur_max"]:
                            direction = "below min" if cur_val < thresh["cur_min"] else "above max"
                            alerts_to_insert.append((
                                deviceID, "CURRENT", phase_num,
                                f"Phase {phase_name}: Current {cur_val}A is {direction} "
                                f"(Threshold: {thresh['cur_min']}–{thresh['cur_max']}A)",
                                ts
                            ))

                if alerts_to_insert:
                    send_alert(alerts_to_insert)
                else:
                    print("[DEBUG] No alerts to insert 🚫")

        send_runtime()			  
        socketio.emit("device_status", {
            "device_id": deviceID,
            "status": "online"
        })

        socketio.emit("live_meter", {
            "deviceID": deviceID,
            "cur1": cur1, "cur2": cur2, "cur3": cur3,
            "vol1": vol1, "vol2": vol2, "vol3": vol3,
            "pow1": pow1, "pow2": pow2, "pow3": pow3,
            "energy1": energy1, "energy2": energy2, "energy3": energy3,
            "freq1": freq1, "freq2": freq2, "freq3": freq3,
            "pf1": pf1, "pf2": pf2, "pf3": pf3,
            "time": time_label,
            "timestamp": ts
        })
        print("✅ Meter data saved & emitted")

    except Exception as e:
        print(f"❌ MQTT Error: {e}")
        import traceback
        traceback.print_exc()

def send_alert(alerts_to_insert):
    if not alerts_to_insert:
        return

    print(f"[DEBUG] Total alerts: {len(alerts_to_insert)}")
    print(f"[DEBUG] Alerts Data: {alerts_to_insert}")

    limit_type = alerts_to_insert[0][1]

    try:
        with sqlite3.connect(DB_FILE) as conn:
            cur = conn.cursor()

            cur.executemany("""
                INSERT INTO alerts (deviceID, alert_type, phase, message, timestamp)
                VALUES (?, ?, ?, ?, ?)
            """, alerts_to_insert)

            message = str(alerts_to_insert[0][3]).lower()

            if "max" in message:
                limit_type = "max"
            elif "min" in message:
                limit_type = "min"
            else:
                limit_type = "other"

            conn.commit()

        print("[DEBUG] Alerts inserted successfully into DB ✅")

    except Exception as e:
        print(f"[ERROR] DB Insert Failed ❌: {e}")

    # Emit alerts to frontend
    for a in alerts_to_insert:
        payload = {
            "deviceID": a[0],
            "alert_type": a[1],
            "limit_type": limit_type,
            "phase": a[2],
            "message": a[3],
            "timestamp": a[4]
        }

        print(f"[DEBUG] Emitting alert: {payload}")

        # Broadcast to all clients
        socketio.emit("alert", payload)           

def send_runtime():

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    query = """
        SELECT COUNT(*) 
        FROM meter_data
        WHERE DATE(inserttimestamp) = DATE('now')
        AND (pow1 > 0 AND pow2 > 0 AND pow3 > 0);
    """

    cursor.execute(query)
    result = cursor.fetchone()
    conn.close()

    minutes = result[0] if result[0] else 0

    hr = minutes // 60
    min = minutes % 60

    runtime = f"{hr:02}:{min:02}"

    socketio.emit("runtime_update", {"runtime": runtime})

# registerr 
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form["name"]
        email = request.form["email"]
        password = generate_password_hash(request.form["password"])

        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()

        try:
            cur.execute(
                "INSERT INTO users (name, email, password) VALUES (?, ?, ?)",
                (name, email, password)
            )
            conn.commit()
        except:
            return "Email already exists"
        finally:
            conn.close()

        return redirect("/login")

    return render_template("register.html")

# login route
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]

        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()

        cur.execute("SELECT id, password FROM users WHERE email=?", (email,))
        user = cur.fetchone()
        conn.close()

        if user and check_password_hash(user[1], password):
            session["user_id"] = user[0]   # ✅ LOGIN
            return redirect("/")
        else:
            return "Invalid email or password"

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

# ---------------- FLASK ROUTE ----------------
@app.route("/")
def index():

    print("🌐 [DEBUG] Index route called")

    if "user_id" not in session:
        print("⚠ [DEBUG] User not logged in, redirecting to login")
        return redirect("/login")

    try:
        print("🟡 [DEBUG] Connecting to database:", DB_FILE)

        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()

        print("🟢 [DEBUG] DB Connected")

        # -------- Fetch latest meter data --------
        print("🟡 [DEBUG] Fetching latest meter data for:", TARGET_DEVICE)

        cur.execute("""
            SELECT *
            FROM meter_data
            WHERE deviceID = ?
            ORDER BY id DESC
            LIMIT 1
        """, (TARGET_DEVICE,))

        row = cur.fetchone()

        print("🟡 [DEBUG] Raw row from DB:", row)

        if row:

            print("🟢 [DEBUG] Meter data found")

            (
                id,
                deviceID,

                vol1, cur1, pow1, energy1, freq1, pf1,
                vol2, cur2, pow2, energy2, freq2, pf2,
                vol3, cur3, pow3, energy3, freq3, pf3,
                load_status,

                insert_ts
            ) = row

            print(f"📊 [DEBUG] Device: {deviceID}")
            print(f"📊 [DEBUG] Voltage: {vol1}, {vol2}, {vol3}")
            print(f"📊 [DEBUG] Current: {cur1}, {cur2}, {cur3}")
            print(f"📊 [DEBUG] Power: {pow1}, {pow2}, {pow3}")
            print(f"📊 [DEBUG] Load Status: {load_status}")

        else:
            print("📭 [DEBUG] No meter data found")

            deviceID = TARGET_DEVICE

            vol1 = vol2 = vol3 = 0
            cur1 = cur2 = cur3 = 0
            pow1 = pow2 = pow3 = 0

            energy1 = energy2 = energy3 = 0
            freq1 = freq2 = freq3 = 0
            pf1 = pf2 = pf3 = 0
             
            load_status = 0    # ✅ default OFF
            


            insert_ts = "N/A"

        # -------- Fetch device status --------
        print("🟡 [DEBUG] Fetching device status")

        cur.execute("""
            SELECT status, last_seen
            FROM device_status
            WHERE device_id = ?
            ORDER BY last_seen DESC
            LIMIT 1
        """, (TARGET_DEVICE,))

        status_data = cur.fetchone()

        print("🟡 [DEBUG] Status row:", status_data)

        if status_data:
            device_status, last_seen = status_data
        else:
            device_status = "offline"
            last_seen = "N/A"

        print(f"📡 [DEBUG] Device Status: {device_status} | Last Seen: {last_seen}")


        print("🟢 [DEBUG] DB Connection Closed")

        print("🚀 [DEBUG] Rendering HTML template")

        # cur.execute("""
        #     SELECT *
        #     FROM alerts
        #     WHERE deviceID = ?
        #     ORDER BY id DESC
        # """, (TARGET_DEVICE,))

        # rows = cur.fetchall()

        # grouped = {}

        # for r in rows:

        #     key = (r[2], r[5])  # alert_type + time

        #     phase = r[4].split(":")[0]

        #     if key not in grouped:
        #         grouped[key] = {
        #             "id": r[0],
        #             "device": r[1],
        #             "type": r[2],
        #             "message": r[2],
        #             "phases": [],
        #             "time": r[5]
        #         }

        #     grouped[key]["phases"].append(phase)

        # alerts = list(grouped.values())

        cur.execute("""
            SELECT *
            FROM alerts
            WHERE deviceID = ?
            ORDER BY id DESC
        """, (TARGET_DEVICE,))

        rows = cur.fetchall()

        grouped = {}

        for r in rows:

            message = str(r[4]).lower()   # force convert to string
            phase = r[4].split(":")[0]

            if "max" in message:
                msg_type = "max"
            elif "min" in message:
                msg_type = "min"
            else:
                msg_type = "other"

            # key me msg_type bhi add karo
            key = (r[2], r[5], msg_type)

            if key not in grouped:
                grouped[key] = {
                    "id": r[0],
                    "device": r[1],
                    "type": r[2],
                    "message": r[3],
                    "phases": [],
                    "time": r[5],
                    "limit_type": msg_type
                }

            grouped[key]["phases"].append(phase)

        alerts = list(grouped.values())

        print(f"⚠ [DEBUG] Alerts fetched-->: {len(alerts)}, alerts: {alerts}")
        


        # ════════════════════════════════════════════════════════
        # ✅ NEW — Fetch latest saved Time & Date
        # ════════════════════════════════════════════════════════
        cur.execute("""
            SELECT hour, minute, day, month, year, set_at
            FROM device_time_settings
            WHERE deviceID = ?
            ORDER BY id DESC
            LIMIT 1
        """, (TARGET_DEVICE,))

        time_row = cur.fetchone()

        if time_row:
            saved_hour   = time_row[0]
            saved_minute = time_row[1]
            saved_day    = time_row[2]
            saved_month  = time_row[3]
            saved_year   = time_row[4]
            time_set_at  = time_row[5]
        else:
            saved_hour = saved_minute = 0
            saved_day  = saved_month  = 1
            saved_year = 2026
            time_set_at = "N/A"

        print(f"🕐 [DEBUG] Saved Time: {saved_hour}:{saved_minute} | {saved_day}/{saved_month}/{saved_year}")

        # ════════════════════════════════════════════════════════
        # ✅ NEW — Fetch latest Scheduling
        # ════════════════════════════════════════════════════════
        cur.execute("""
            SELECT is_enabled, on_hour, on_minute, off_hour, off_minute, set_at
            FROM device_scheduling
            WHERE deviceID = ?
            ORDER BY id DESC
            LIMIT 1
        """, (TARGET_DEVICE,))

        sched_row = cur.fetchone()

        if sched_row:
            sched_enabled    = sched_row[0]
            sched_on_hour    = sched_row[1]
            sched_on_minute  = sched_row[2]
            sched_off_hour   = sched_row[3]
            sched_off_minute = sched_row[4]
            sched_set_at     = sched_row[5]
        else:
            sched_enabled    = 0
            sched_on_hour    = 0
            sched_on_minute  = 0
            sched_off_hour   = 0
            sched_off_minute = 0
            sched_set_at     = "N/A"

        print(f"📅 [DEBUG] Scheduling: enabled={sched_enabled} | ON {sched_on_hour}:{sched_on_minute} OFF {sched_off_hour}:{sched_off_minute}")

        conn.close()
        print("🟢 [DEBUG] DB Connection Closed")
        print("🚀 [DEBUG] Rendering HTML template")


        conn.close()

        return render_template(
            "index.html",

            deviceID=deviceID,

            # Phase 1
					  
            vol1=vol1,
            cur1=cur1,
					  
            pow1=pow1,
            energy1=energy1,
            freq1=freq1,
            pf1=pf1,

            # Phase 2
            vol2=vol2,
            cur2=cur2,
            pow2=pow2,
            energy2=energy2,
            freq2=freq2,
            pf2=pf2,

            # Phase 3
            vol3=vol3,
            cur3=cur3,
            pow3=pow3,
            energy3=energy3,
            freq3=freq3,
            pf3=pf3,
            load_status=load_status,     # ✅ passed to template

            timestamp=insert_ts,

            device_status=device_status,
            last_seen=last_seen,

            alerts=alerts,


            # ✅ NEW — Time & Date
            saved_hour=saved_hour,
            saved_minute=saved_minute,
            saved_day=saved_day,
            saved_month=saved_month,
            saved_year=saved_year,
            time_set_at=time_set_at,

            # ✅ NEW — Scheduling
            sched_enabled=sched_enabled,
            sched_on_hour=sched_on_hour,
            sched_on_minute=sched_on_minute,
            sched_off_hour=sched_off_hour,
            sched_off_minute=sched_off_minute,
            sched_set_at=sched_set_at,


        )

    except Exception as e:

        print("❌ [ERROR] Exception occurred:", e)

        import traceback
        traceback.print_exc()

        return render_template(
            "index.html",
            deviceID=TARGET_DEVICE,
            vol1=0, cur1=0, pow1=0, energy1=0, freq1=0, pf1=0,
            vol2=0, cur2=0, pow2=0, energy2=0, freq2=0, pf2=0,
            vol3=0, cur3=0, pow3=0, energy3=0, freq3=0, pf3=0,
            load_status=0,               # ✅ default in except
            timestamp="N/A",
            device_status="offline",
            last_seen="N/A",
            alerts=[],
             # ✅ NEW — defaults in except
            saved_hour=0, saved_minute=0,
            saved_day=1,  saved_month=1, saved_year=2026,
            time_set_at="N/A",
            sched_enabled=0,
            sched_on_hour=0,  sched_on_minute=0,
            sched_off_hour=0, sched_off_minute=0,
            sched_set_at="N/A",
        )


# ---------------- EXCEL DOWNLOAD ROUTE ----------------
@app.route("/download_excel")
def download_excel():
    """Download today's meter data as Excel file"""
    try:
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()

        today = datetime.date.today().strftime("%Y-%m-%d")
        
        selected_date = request.args.get("date")
        
        if selected_date:
           query_date = selected_date
        else:
           query_date = today

        cur.execute("""
            SELECT deviceID, cur1, cur2, cur3, vol1, vol2, vol3, pow1, pow2, pow3 , energy1 , energy2 , energy3 , freq1 , freq2 , freq3 , pf1 , pf2 , pf3,  inserttimestamp
            FROM meter_data
            WHERE date(inserttimestamp) = ?
            ORDER BY inserttimestamp ASC
        """, (query_date,))

        rows = cur.fetchall()
        conn.close()

        wb = Workbook()
        sheet = wb.active
        sheet.title = f"Meter Data {today}"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = ["Device ID", "Current 1 (A)", "Current 2 (A)", "Current 3 (A)", 
                   "Voltage 1 (V)", "Voltage 2 (V)", "Voltage 3 (V)",
                   "Power 1 (W)", "Power 2 (W)", "Power 3 (W)", 
                   "Energy 1  (kWh)", "Energy 2 (kWh)" , "Energy 3 (kWh)" , 
                   "Frequency 1 (Hz) " , "Frequency 2 (Hz)" , "Frequency 3 (Hz)" ,
                   "Power Factor 1 " , "Power Factor 2" , "Power Factor 3" ,"Timestamp"]
        
        sheet.append(headers)
        
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for row in rows:
            sheet.append(row)
            for cell in sheet[sheet.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

        column_widths = [15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 20]
        for idx, width in enumerate(column_widths, 1):
            sheet.column_dimensions[chr(64 + idx)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"meter_data_{today}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"❌ Excel Download Error: {e}")
        import traceback
        traceback.print_exc()
        return "Error generating Excel file", 500
    
# ── Save Threshold ──────────────────────────────────────────────────────────
@app.route("/save_threshold", methods=["POST"])
def save_threshold():
    data = request.get_json()
    device_id  = data.get("device_id")
    ttype      = data.get("type")       # "voltage" or "current"
    t_min      = data.get("min")
    t_max      = data.get("max")

    if not all([device_id, ttype, t_min is not None, t_max is not None]):
        return jsonify({"error": "Missing fields"}), 400

    if float(t_min) >= float(t_max):
        return jsonify({"error": "Min must be less than Max"}), 400

    now = datetime.datetime.now().isoformat()

    with sqlite3.connect(DB_FILE) as conn:
        cur = conn.cursor()

        # Make sure row exists first
        cur.execute("""
            INSERT OR IGNORE INTO device_thresholds
            (device_id, updated_at) VALUES (?, ?)
        """, (device_id, now))

        if ttype == "voltage":
            cur.execute("""
                UPDATE device_thresholds
                SET vol_min=?, vol_max=?, updated_at=?
                WHERE device_id=?
            """, (t_min, t_max, now, device_id))
        elif ttype == "current":
            cur.execute("""
                UPDATE device_thresholds
                SET cur_min=?, cur_max=?, updated_at=?
                WHERE device_id=?
            """, (t_min, t_max, now, device_id))
        else:
            return jsonify({"error": "Invalid type"}), 400

        conn.commit()

    return jsonify({"success": True, "message": f"{ttype.capitalize()} threshold saved!"})


# ── Get Threshold ───────────────────────────────────────────────────────────
@app.route("/get_thresholds/<device_id>")
def get_thresholds(device_id):
    with sqlite3.connect(DB_FILE) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM device_thresholds WHERE device_id=?", (device_id,)
        ).fetchone()
    return jsonify(dict(row) if row else {})




# ---------------- SOCKET EVENTS ----------------
@socketio.on("get_today_graph")
def send_today_graph():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()

    today = datetime.date.today().strftime("%Y-%m-%d")

    # Line graph data (existing)
    cur.execute("""
        SELECT
            strftime('%H:%M', inserttimestamp) AS time,
            AVG(cur1), AVG(cur2), AVG(cur3), AVG(vol1), AVG(vol2), AVG(vol3), AVG(pow1), AVG(pow2), AVG(pow3),
             
        AVG(energy1), AVG(energy2), AVG(energy3),
        AVG(freq1), AVG(freq2), AVG(freq3),
        AVG(pf1), AVG(pf2), AVG(pf3)   
                
        FROM meter_data
        WHERE date(inserttimestamp) = ?
        GROUP BY strftime('%H:%M', inserttimestamp)
        ORDER BY inserttimestamp ASC
    """, (today,))

    rows = cur.fetchall()
    
    # Pie chart data - latest values
    cur.execute("""
        SELECT cur1, cur2, cur3, vol1, vol2, vol3, pow1, pow2, pow3, energy1, energy2, energy3, freq1, freq2, freq3, pf1, pf2, pf3
        FROM meter_data
        WHERE date(inserttimestamp) = ?
        ORDER BY inserttimestamp DESC
        LIMIT 1
    """, (today,))
    
    latest = cur.fetchone()
    conn.close()

    
    labels = []

    c1 = []
    c2 = []
    c3 = []

    v1 = []
    v2 = []
    v3 = []

    p1 = []
    p2 = []
    p3 = []

    e1 = []
    e2 = []
    e3 = []

    f1 = []
    f2 = []
    f3 = []

    pf1 = []
    pf2 = []
    pf3 = []


    for r in rows:

        labels.append(r[0])

        c1.append(r[1])
        c2.append(r[2])
        c3.append(r[3])

        v1.append(r[4])
        v2.append(r[5])
        v3.append(r[6])

        p1.append(r[7])
        p2.append(r[8])
        p3.append(r[9])

        e1.append(r[10])
        e2.append(r[11])
        e3.append(r[12])

        f1.append(r[13])
        f2.append(r[14])
        f3.append(r[15])

        pf1.append(r[16])
        pf2.append(r[17])
        pf3.append(r[18])


    # ---------------- PIE DATA ----------------

    pie_data = {

        "current": {
            "r": latest[0] if latest else 0,
            "y": latest[1] if latest else 0,
            "b": latest[2] if latest else 0
        },

        "voltage": {
            "r": latest[3] if latest else 0,
            "y": latest[4] if latest else 0,
            "b": latest[5] if latest else 0
        },

        "power": {
            "r": latest[6] if latest else 0,
            "y": latest[7] if latest else 0,
            "b": latest[8] if latest else 0
        },

        "energy": {
            "r": latest[9] if latest else 0,
            "y": latest[10] if latest else 0,
            "b": latest[11] if latest else 0
        },

        "frequency": {
            "r": latest[12] if latest else 0,
            "y": latest[13] if latest else 0,
            "b": latest[14] if latest else 0
        },

        "pf": {
            "r": latest[15] if latest else 0,
            "y": latest[16] if latest else 0,
            "b": latest[17] if latest else 0
        }

    }

   
# ---------------- EMIT TO DASHBOARD ----------------

    socketio.emit("today_graph_data", {

        "labels": labels,

        "cur1": c1,
        "cur2": c2,
        "cur3": c3,

        "vol1": v1,
        "vol2": v2,
        "vol3": v3,

        "pow1": p1,
        "pow2": p2,
        "pow3": p3,

        "energy1": e1,
        "energy2": e2,
        "energy3": e3,

        "freq1": f1,
        "freq2": f2,
        "freq3": f3,

        "pf1": pf1,
        "pf2": pf2,
        "pf3": pf3,

        "pie": pie_data
    }, room=request.sid)



# ---------------- CUSTOM RANGE GRAPH ----------------
@socketio.on("get_custom_graph")
def send_custom_graph(data):

    start = data.get("start")
    end = data.get("end")

    print("📊 Custom Graph Request")
    print("Start:", start)
    print("End:", end)

    try:

        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()

        # -------- FETCH GRAPH DATA --------
        cur.execute("""
            SELECT
                strftime('%H:%M', inserttimestamp) AS time,

                AVG(cur1), AVG(cur2), AVG(cur3),
                AVG(vol1), AVG(vol2), AVG(vol3),
                AVG(pow1), AVG(pow2), AVG(pow3),

                AVG(energy1), AVG(energy2), AVG(energy3),
                AVG(freq1), AVG(freq2), AVG(freq3),
                AVG(pf1), AVG(pf2), AVG(pf3)

            FROM meter_data

            WHERE date(inserttimestamp) BETWEEN ? AND ?
            AND deviceID = ?
									
					  

            GROUP BY time
            ORDER BY time ASC
        """, (start, end, TARGET_DEVICE))

        rows = cur.fetchall()

        conn.close()

        print("📈 Rows fetched:", len(rows))

        # -------- GRAPH ARRAYS --------
        labels = []

        c1, c2, c3 = [], [], []
        v1, v2, v3 = [], [], []
        p1, p2, p3 = [], [], []

        e1, e2, e3 = [], [], []
        f1, f2, f3 = [], [], []

        pf1, pf2, pf3 = [], [], []

        for r in rows:

            labels.append(r[0])

            c1.append(r[1])
            c2.append(r[2])
            c3.append(r[3])

            v1.append(r[4])
            v2.append(r[5])
            v3.append(r[6])

            p1.append(r[7])
            p2.append(r[8])
            p3.append(r[9])

            e1.append(r[10])
            e2.append(r[11])
            e3.append(r[12])

            f1.append(r[13])
            f2.append(r[14])
            f3.append(r[15])

            pf1.append(r[16])
            pf2.append(r[17])
            pf3.append(r[18])

        # -------- EMIT DATA --------
        socketio.emit("custom_graph_data", {

            "labels": labels,

            "cur1": c1,
            "cur2": c2,
            "cur3": c3,

            "vol1": v1,
            "vol2": v2,
            "vol3": v3,

            "pow1": p1,
            "pow2": p2,
            "pow3": p3,

            "energy1": e1,
            "energy2": e2,
            "energy3": e3,

            "freq1": f1,
            "freq2": f2,
            "freq3": f3,

            "pf1": pf1,
            "pf2": pf2,
            "pf3": pf3
        })

        print("🚀 Custom graph data sent")

    except Exception as e:

        print("❌ Custom Graph Error:", e)

        import traceback
        traceback.print_exc()

# ---------------- SOCKETIO EVENTS ----------------
@socketio.on('connect')
def on_socket_connect():
    """
    Jab bhi client connect/refresh kare, latest data emit karo
    """
    print(f"🔌 Client connected: {request.sid}")
    print("=" * 80)
    send_runtime()
    try:
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()

        # ---------------- FETCH LATEST METER DATA ----------------
        print(f"📊 Fetching latest meter data for device: {TARGET_DEVICE}")

        cur.execute("""
            SELECT deviceID,
                   cur1, cur2, cur3,
                   vol1, vol2, vol3,
                   pow1, pow2, pow3,
                   energy1, energy2, energy3,
                   freq1, freq2, freq3,
                   pf1, pf2, pf3,
                   inserttimestamp
            FROM meter_data
            WHERE deviceID = ? AND date(inserttimestamp) = date('now')
            ORDER BY id DESC LIMIT 1
        """, (TARGET_DEVICE,))

        latest_data = cur.fetchone()

        if latest_data:

            (deviceID,
             cur1, cur2, cur3,
             vol1, vol2, vol3,
             pow1, pow2, pow3,
             energy1, energy2, energy3,
             freq1, freq2, freq3,
             pf1, pf2, pf3,
             timestamp) = latest_data

            print(
                f"✅ Data | Cur: {cur1},{cur2},{cur3} | "
                f"Vol: {vol1},{vol2},{vol3} | "
                f"Pow: {pow1},{pow2},{pow3} | "
                f"Energy: {energy1},{energy2},{energy3} | "
                f"Freq: {freq1},{freq2},{freq3} | "
                f"PF: {pf1},{pf2},{pf3} | "
                f"Time: {timestamp}"
            )

            # Convert timestamp
            db_time = datetime.datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
												
            # ---------------- EMIT DATA ----------------
													 
            socketio.emit("live_current", {

                "deviceID": deviceID,

                # current
                "cur1": cur1,
                "cur2": cur2,
                "cur3": cur3,

                # voltage
                "vol1": vol1,
                "vol2": vol2,
                "vol3": vol3,

                # power
                "pow1": pow1,
                "pow2": pow2,
                "pow3": pow3,

                # new values
                "energy1": energy1,
                "energy2": energy2,
                "energy3": energy3,
                "frequency1": freq1,
                "frequency2": freq2,
                "frequency3": freq3,
                "power_factor1": pf1,
                "power_factor2": pf2,
                "power_factor3": pf3,

                "time": db_time.strftime("%H:%M"),
                "timestamp": timestamp				  
            }, room=request.sid)

            print(f"📤 Emitted meter data to client {request.sid}")

        else:
            print("📭 No meter data found")

            socketio.emit("live_current", {

                "deviceID": TARGET_DEVICE,

                "cur1": 0,
                "cur2": 0,
                "cur3": 0,

                "vol1": 0,
                "vol2": 0,
                "vol3": 0,

                "pow1": 0,
                "pow2": 0,
                "pow3": 0,

                "energy1": 0,
                "energy2": 0,
                "energy3": 0,
                "frequency1": 0,
                "frequency2": 0,
                "frequency3": 0,
                "power_factor1": 0,
                "power_factor2": 0,
                "power_factor3": 0,

                "time": datetime.datetime.now().strftime("%H:%M"),
                "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            }, room=request.sid)
																		 

        # ---------------- DEVICE STATUS ----------------
        cur.execute("""
            SELECT status, last_seen
            FROM device_status
            WHERE device_id = ?
            ORDER BY last_seen DESC LIMIT 1
        """, (TARGET_DEVICE,))

        status_data = cur.fetchone()

        status, last_seen = status_data if status_data else ("offline", "N/A")

        socketio.emit("device_status", {
            "device_id": TARGET_DEVICE,
            "status": status
        }, room=request.sid)
																			   

        print(f"📤 Emitted device_status ({status})")

        conn.close()

        print("=" * 80)

    except Exception as e:
        print(f"❌ Error in on_socket_connect: {e}")

        import traceback
        traceback.print_exc()


@socketio.on("handle_on_off")
def handle_on_off(data):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    try:
        device_id = data.get("device")       # "3Phase862733079332931"
        intensity = data.get("intensity")    # 100 = ON, 0 = OFF
        print("Received on/off request:", data)

        logging.info(f"🔌 SocketIO On/Off — Device: {device_id}, Intensity: {intensity}")

        load_status = 1 if int(intensity) > 0 else 0

        # ── 1. Check device exists ───────────────────────────────────────
        cursor.execute("""
            SELECT id, deviceID
            FROM meter_data
            WHERE deviceID = ?
            ORDER BY id DESC
            LIMIT 1
        """, (device_id,))

        row = cursor.fetchone()

        if not row:
            logging.error(f"❌ Device not found: {device_id}")
            emit("error", {"message": f"Device {device_id} not found"})
            return
                
                # ── 2. MQTT Publish ────────────────────────────
        topic = f"{device_id}/control"

        # Command string
        command = f"lte:Load:{load_status}"

        # ✅ Send plain string (IMPORTANT FIX)
        payload = command

        print(f"Publishing → Topic: {topic}")
        print(f"Payload: {payload}")

        logging.info(f"📡 Publishing MQTT → Topic: {topic} | Payload: {payload}")

        try:
            mqttc.publish(topic, payload)
            logging.info("✅ MQTT Publish successful")
        except Exception as e:
            logging.error(f"❌ MQTT Publish failed: {e}")
            emit("error", {"message": f"MQTT publish failed: {str(e)}"})
            return

        # ── 3. Update load_status in DB ──────────────────────────────────
        cursor.execute("""
        UPDATE meter_data
        SET load_status = ?
        WHERE id = (
            SELECT id FROM meter_data
            WHERE deviceID = ?
            ORDER BY id DESC
            LIMIT 1
        )
    """, (load_status, device_id))

        conn.commit()
        logging.info(f"✅ DB Updated — deviceID: {device_id}, load_status: {load_status}")

        # ── 4. Confirm back to frontend ──────────────────────────────────
        emit("on_off_response", {
            "device_id":   device_id,
            "load_status": load_status,
            "message":     f"Load is {'ON' if load_status else 'OFF'}"
        })

    except sqlite3.Error as e:
        logging.error(f"❌ DB Error: {e}")
        emit("error", {"message": str(e)})

    finally:
        conn.close()

# ─────────────────────────────────────────────────────────────────────────────
# SOCKETIO  —  SET TIME & DATE
# Command: lte:setTime:Hour:Minute:Day:Month:Year
# ─────────────────────────────────────────────────────────────────────────────
@socketio.on("set_time_date")
def handle_set_time_date(data):
    """
    Expects:
    {
        device: "3Phase862733079332931",
        hour:   12,
        minute: 10,
        day:    31,
        month:  3,
        year:   2026
    }
    """
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    try:
        device_id = data.get("device")
        hour      = int(data.get("hour",   0))
        minute    = int(data.get("minute", 0))
        day       = int(data.get("day",    1))
        month     = int(data.get("month",  1))
        year      = int(data.get("year",   2026))

        logging.info(f"🕐 Set Time — Device: {device_id} | {hour}:{minute} {day}/{month}/{year}")

        # ── Validate ──────────────────────────────────────────────────────────
        errors = []

        if not (0 <= hour   <= 23): errors.append("Hour must be 0–23")
        if not (0 <= minute <= 59): errors.append("Minute must be 0–59")
        if not (1 <= day    <= 31): errors.append("Day must be 1–31")
        if not (1 <= month  <= 12): errors.append("Month must be 1–12")
        if not (2000 <= year <= 2099): errors.append("Year must be 2000–2099")

        if errors:
            emit("set_time_error", {"message": " | ".join(errors)})
            return

        # ── Build MQTT command ────────────────────────────────

        topic = f"{device_id}/control"

        payload = f"lte:setTime:{hour}:{minute}:{day}:{month}:{year}"

        print(f"Publishing → Topic: {topic}")
        print(f"Payload: {payload}")

        logging.info(f"📡 MQTT → Topic: {topic}")

        try:
            mqttc.publish(topic, payload)
            logging.info("✅ MQTT Publish successful")
        except Exception as e:
            logging.error(f"❌ MQTT Publish failed: {e}")
            emit("set_time_error", {"message": f"MQTT failed: {str(e)}"})
            return

        # ── Save to DB ────────────────────────────────────────────────────────
        cursor.execute("""
            INSERT INTO device_time_settings
                (deviceID, hour, minute, day, month, year)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (device_id, hour, minute, day, month, year))

        conn.commit()
        logging.info(f"✅ Time saved to DB — {device_id}")

        # ── Confirm to frontend ───────────────────────────────────────────────
        emit("set_time_response", {
            "device_id": device_id,
            "message":   f"Date And Time Updated → {hour:02d}:{minute:02d} | {day:02d}/{month:02d}/{year}",
            "hour":      hour,
            "minute":    minute,
            "day":       day,
            "month":     month,
            "year":      year
        })

    except Exception as e:
        logging.error(f"❌ set_time_date error: {e}")
        emit("set_time_error", {"message": str(e)})
    finally:
        conn.close()

# ─────────────────────────────────────────────────────────────────────────────
# SOCKETIO  —  SET SCHEDULING
# ON:  lte:SchedulingStatus:ONHour:ONMinute:OFFHour:OFFMinute  →  lte:1:10:30:12:30
# OFF: lte:0
# ─────────────────────────────────────────────────────────────────────────────
@socketio.on("set_scheduling")
def handle_set_scheduling(data):
    """
    Expects:
    {
        device:     "3Phase862733079332931",
        is_enabled: 1,           -- 1=ON, 0=OFF
        on_hour:    10,          -- required only when is_enabled = 1
        on_minute:  30,
        off_hour:   12,
        off_minute: 30
    }
    """
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    try:
        device_id  = data.get("device")
        is_enabled = int(data.get("is_enabled", 0))

        logging.info(f"📅 Scheduling — Device: {device_id} | Enabled: {is_enabled}")

        if is_enabled == 1:
            on_hour    = int(data.get("on_hour",    0))
            on_minute  = int(data.get("on_minute",  0))
            off_hour   = int(data.get("off_hour",   0))
            off_minute = int(data.get("off_minute", 0))

            # ── Validate ──────────────────────────────────────────────────────
            errors = []

            if not (0 <= on_hour    <= 23): errors.append("ON Hour must be 0–23")
            if not (0 <= on_minute  <= 59): errors.append("ON Minute must be 0–59")
            if not (0 <= off_hour   <= 23): errors.append("OFF Hour must be 0–23")
            if not (0 <= off_minute <= 59): errors.append("OFF Minute must be 0–59")

            if errors:
                emit("scheduling_error", {"message": " | ".join(errors)})
                return
            
            # ── Scheduling ON Command ────────────────────────────

            topic = f"{device_id}/control"

            # ✅ CORRECT PAYLOAD (FINAL FIX)
            payload = f"lte:schedulingDevice:1:{on_hour}:{on_minute}:{off_hour}:{off_minute}"

            print(f"Publishing → Topic: {topic}")
            print(f"Payload: {payload}")

            logging.info(f"📡 MQTT Scheduling ON → {payload}")

        else:
            topic = f"{device_id}/control"

            # lte:0  (turn off scheduling)
       
            # ✅ CORRECT OFF PAYLOAD
            payload = "lte:schedulingDevice:0"

            logging.info(f"📡 MQTT Scheduling OFF → {topic}")
            logging.info(f"📡 MQTT Scheduling OFF → {payload}")

        # ── Publish MQTT ──────────────────────────────────────────────────────
        try:
            mqttc.publish(topic, payload)
            logging.info("✅ MQTT Publish successful")
        except Exception as e:
            logging.error(f"❌ MQTT Publish failed: {e}")
            emit("scheduling_error", {"message": f"MQTT failed: {str(e)}"})
            return

        # ── Save to DB ────────────────────────────────────────────────────────
        cursor.execute("""
            INSERT INTO device_scheduling
                (deviceID, is_enabled, on_hour, on_minute, off_hour, off_minute)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (device_id, is_enabled, on_hour, on_minute, off_hour, off_minute))

        conn.commit()
        logging.info(f"✅ Scheduling saved to DB — {device_id}")

        # ── Confirm to frontend ───────────────────────────────────────────────
        if is_enabled:
            msg = (f"Scheduling Time Updated → "
                   f"ON {on_hour:02d}:{on_minute:02d} | OFF {off_hour:02d}:{off_minute:02d}")
        else:
            msg = "Scheduling Disabled"

        emit("scheduling_response", {
            "device_id":  device_id,
            "is_enabled": is_enabled,
            "on_hour":    on_hour,
            "on_minute":  on_minute,
            "off_hour":   off_hour,
            "off_minute": off_minute,
            "message":    msg
        })

    except Exception as e:
        logging.error(f"❌ set_scheduling error: {e}")
        emit("scheduling_error", {"message": str(e)})
    finally:
        conn.close()




@app.route("/delete_alert", methods=["POST"])
def delete_alert():
    try:
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()

        data = request.json

        print("alert comes to delte:", data)

        device = data["device"]
        alert_type = data["type"]
        timestamp = data["time"]

        cur = conn.cursor()

        cur.execute("""
            DELETE FROM alerts
            WHERE deviceID = ?
            AND alert_type = ?
            AND timestamp = ?
        """,(device, alert_type, timestamp))

        conn.commit()

        return {"status":"deleted"}
    
    except Exception as e:
        print(f"❌ Error deleting alert: {e}")
        return {"status":"error", "message": str(e)}, 500
    
    finally:
        conn.close()

# ---------------- MQTT INIT ----------------
mqttc.username_pw_set(USERNAME, PASSWORD)
mqttc.on_connect = on_connect
mqttc.on_disconnect = on_disconnect
mqttc.on_message = on_message

def mqtt_connect():
    mqttc.connect(broker, port, keepalive=60)
    mqttc.loop_start()

mqtt_connect()

def initialize_device_status():
    now = datetime.datetime.now()
    ts = now.strftime('%Y-%m-%d %H:%M:%S')

    with sqlite3.connect(DB_FILE) as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT OR REPLACE INTO device_status (device_id, status, last_seen)
            VALUES (?, ?, ?)
        """, (TARGET_DEVICE, "offline", ts))
        conn.commit()

    print(f"🔴 Device {TARGET_DEVICE} set to OFFLINE on startup")

initialize_device_status()


# ---------------- RUN APP ----------------
if __name__ == "__main__":
    socketio.run(app, host="0.0.0.0", port=5001)